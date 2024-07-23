import openpyxl

def main():
  
    counting_sort_times = {
        "15K": 0.527,
        "500K": 14.183
    }

    insertion_sort_times = {
        "15K": 0.068,
        "500K": 101.876
    }

  
    squares = {
        "15K": 15000 ** 2,
        "500K": 500000 ** 2
    }


    counting_sort_ratios = {
        "15K": counting_sort_times["15K"] / counting_sort_times["500K"],
        "500K": 1
    }

    insertion_sort_ratios = {
        "15K": insertion_sort_times["15K"] / insertion_sort_times["500K"],
        "500K": 1
    }

    workbook = openpyxl.Workbook()

    worksheet1 = workbook.active
    worksheet1.title = "Counting Sort Times"
    worksheet1.cell(row=1, column=1, value='Array Size')
    worksheet1.cell(row=1, column=2, value='Time (s)')
    worksheet1.cell(row=1, column=3, value='Square')
    worksheet1.cell(row=1, column=4, value='Ratio')

    row = 2
    for size, time in counting_sort_times.items():
        worksheet1.cell(row=row, column=1, value=size)
        worksheet1.cell(row=row, column=2, value=time)
        worksheet1.cell(row=row, column=3, value=squares[size])
        worksheet1.cell(row=row, column=4, value=counting_sort_ratios[size])
        row += 1

    worksheet2 = workbook.create_sheet(title="Insertion Sort Times")
    worksheet2.cell(row=1, column=1, value='Array Size')
    worksheet2.cell(row=1, column=2, value='Time (s)')
    worksheet2.cell(row=1, column=3, value='Square')
    worksheet2.cell(row=1, column=4, value='Ratio')

   
    row = 2
    for size, time in insertion_sort_times.items():
        worksheet2.cell(row=row, column=1, value=size)
        worksheet2.cell(row=row, column=2, value=time)
        worksheet2.cell(row=row, column=3, value=squares[size])
        worksheet2.cell(row=row, column=4, value=insertion_sort_ratios[size])
        row += 1

    
    workbook.save(filename="sort_times.xlsx")

if __name__ == "__main__":
    main()
